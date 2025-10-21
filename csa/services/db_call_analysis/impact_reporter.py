"""
영향도 분석 리포트 생성 모듈

Markdown, Excel, JSON 형식의 영향도 분석 리포트 생성
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from csa.models.impact import ImpactAnalysisResult
from csa.diagrams.impact import ImpactMermaidGenerator
from csa.utils.logger import get_logger

logger = get_logger(__name__)


class ImpactReporter:
    """영향도 분석 리포트 생성기

    Markdown, Excel, JSON 형식의 리포트를 생성합니다.
    """

    def generate_markdown(
        self,
        result: ImpactAnalysisResult,
        filepath: Path,
    ) -> bool:
        """Markdown 리포트 생성

        Args:
            result: ImpactAnalysisResult 객체
            filepath: 저장 경로

        Returns:
            성공 여부
        """
        try:
            lines = []

            # 1. 분석 개요
            lines.append("# 영향도 분석 보고서")
            lines.append("")
            lines.append("## 1. 분석 개요")
            lines.append(f"- **분석 대상**: `{result.target_name}`")
            if result.project_name:
                lines.append(f"- **프로젝트**: {result.project_name}")
            else:
                lines.append("- **프로젝트**: 전체")
            lines.append(f"- **분석 유형**: {result.analysis_type}")
            lines.append(f"- **분석 일시**: {result.timestamp}")
            lines.append("")

            # 2. 영향도 요약
            summary = result.summary
            lines.append("## 2. 영향도 요약")
            lines.append("- **총 영향 범위**:")
            lines.append(f"  - 영향받는 클래스: {summary.total_impacted_classes}개")
            lines.append(f"  - 영향받는 메서드: {summary.total_impacted_methods}개")
            lines.append(f"  - 영향받는 패키지: {summary.total_impacted_packages}개")
            lines.append("- **영향 깊이**:")
            lines.append(f"  - 최대 호출 깊이: {summary.max_depth}")
            lines.append(f"  - 평균 호출 깊이: {summary.avg_depth}")

            # 리스크 등급
            risk_dist = summary.risk_distribution
            total_high = risk_dist.get("HIGH", 0)
            total_medium = risk_dist.get("MEDIUM", 0)
            total_low = risk_dist.get("LOW", 0)

            if total_high > 0:
                lines.append("- **리스크 등급**: **HIGH** ⚠️")
            elif total_medium > 0:
                lines.append("- **리스크 등급**: **MEDIUM** ⚠")
            else:
                lines.append("- **리스크 등급**: **LOW** ✓")

            lines.append(f"  - High: {total_high}개 메서드")
            lines.append(f"  - Medium: {total_medium}개 메서드")
            lines.append(f"  - Low: {total_low}개 메서드")
            lines.append("")

            # 3. 영향도 트리 (Level별)
            if result.impact_tree:
                lines.append("## 3. 영향도 트리 (Level별)")
                lines.append("")

                for level in sorted(result.impact_tree.keys()):
                    nodes = result.impact_tree[level]
                    if not nodes:
                        continue

                    if level == 1:
                        lines.append(f"### Level {level} (직접 영향 - Depth 0)")
                    else:
                        max_depth_in_level = max(node.depth for node in nodes)
                        lines.append(f"### Level {level} (간접 영향 - Depth {max_depth_in_level})")

                    for node in nodes:
                        risk_icon = {"HIGH": "🔴", "MEDIUM": "🟡", "LOW": "🟢"}.get(node.risk_grade, "⚪")

                        # 패키지.클래스.메서드
                        full_name = f"{node.package_name}.{node.class_name}.{node.method_name}"

                        # SQL 정보 (있는 경우)
                        if node.sql_id:
                            sql_info = f"({node.sql_type}, Complexity: {node.sql_complexity})"
                        else:
                            sql_info = ""

                        lines.append(f"- {risk_icon} **{full_name}** {sql_info}")

                    lines.append("")
            else:
                lines.append("## 3. 영향도 트리")
                lines.append("영향받는 코드가 없습니다.")
                lines.append("")

            # 4. 패키지별 통계
            if result.package_summary:
                lines.append("## 4. 패키지별 통계")
                lines.append("")
                lines.append("| Package | 클래스 수 | 메서드 수 | 평균 Depth | High Risk | Medium Risk | Low Risk |")
                lines.append("|---------|----------|----------|-----------|-----------|-------------|----------|")

                for pkg in result.package_summary:
                    lines.append(
                        f"| {pkg.package_name} "
                        f"| {pkg.impacted_classes} "
                        f"| {pkg.impacted_methods} "
                        f"| {pkg.avg_depth} "
                        f"| {pkg.risk_distribution.get('HIGH', 0)} "
                        f"| {pkg.risk_distribution.get('MEDIUM', 0)} "
                        f"| {pkg.risk_distribution.get('LOW', 0)} |"
                    )
                lines.append("")

            # 5. SQL 상세 정보 (테이블 분석 시)
            if result.sql_details:
                lines.append("## 5. SQL 상세 정보")
                lines.append("")
                lines.append("| SQL ID | Type | Mapper | Complexity | Query |")
                lines.append("|--------|------|--------|-----------|-------|")

                for sql in result.sql_details:
                    query_preview = sql.query_preview[:50] + "..." if sql.query_preview and len(sql.query_preview) > 50 else sql.query_preview or ""
                    query_preview_md = query_preview.replace("|", "\\|")  # 파이프 이스케이프

                    lines.append(
                        f"| {sql.sql_id} "
                        f"| {sql.sql_type} "
                        f"| {sql.mapper_class}.{sql.mapper_method} "
                        f"| {sql.complexity} "
                        f"| `{query_preview_md}` |"
                    )
                lines.append("")

            # 6. 권장 테스트 범위
            if result.test_scope:
                lines.append("## 6. 권장 테스트 범위")
                lines.append("")

                existing_tests = [item for item in result.test_scope if item.status == "존재"]
                missing_tests = [item for item in result.test_scope if item.status == "미존재"]

                if existing_tests:
                    lines.append("### 기존 테스트 (실행 권장)")
                    for item in existing_tests:
                        lines.append(f"- ✓ `{item.test_class}` (메서드 {item.test_method_count}개)")
                    lines.append("")

                if missing_tests:
                    lines.append("### 테스트 미존재 (작성 권장)")
                    for item in missing_tests:
                        lines.append(f"- ⚠️ `{item.impacted_class}` - 테스트 클래스 없음")
                    lines.append("")

                # 테스트 커버리지
                total_classes = len(result.test_scope)
                tested_classes = len(existing_tests)
                coverage_pct = (tested_classes / total_classes * 100) if total_classes > 0 else 0
                lines.append(f"**테스트 커버리지**: {tested_classes}/{total_classes} ({coverage_pct:.1f}%)")
                lines.append("")

            # 7. 변경 시 주의사항
            lines.append("## 7. 변경 시 주의사항")

            if result.has_circular_reference:
                lines.append(f"- ⚠️ **순환 참조 감지**: {len(result.circular_paths)}개")
                for path in result.circular_paths[:5]:  # 최대 5개만 표시
                    lines.append(f"  - `{path}`")
                if len(result.circular_paths) > 5:
                    lines.append(f"  - ... 외 {len(result.circular_paths) - 5}개")
            else:
                lines.append("- 🔄 **순환 참조**: 없음")

            if total_high > 0:
                lines.append(f"- ⚠️ **High Risk** 메서드 {total_high}개 포함")
                lines.append("  - 변경 시 충분한 테스트 필요")

            # 파일 저장
            with open(filepath, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            logger.info(f"Markdown 리포트 생성 완료: {filepath}")
            return True

        except Exception as exc:
            logger.error(f"Markdown 리포트 생성 실패: {exc}", exc_info=True)
            return False

    def generate_excel(
        self,
        result: ImpactAnalysisResult,
        filepath: Path,
    ) -> bool:
        """Excel 리포트 생성 (다중 시트)

        Args:
            result: ImpactAnalysisResult 객체
            filepath: 저장 경로

        Returns:
            성공 여부
        """
        try:
            wb = Workbook()
            # 기본 시트 제거
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # 시트 생성
            self._generate_summary_sheet(wb, result)
            self._generate_detail_sheet(wb, result)
            self._generate_package_sheet(wb, result)

            if result.sql_details:
                self._generate_sql_sheet(wb, result)

            if result.test_scope:
                self._generate_test_scope_sheet(wb, result)

            # 파일 저장
            wb.save(filepath)
            logger.info(f"Excel 리포트 생성 완료: {filepath}")
            return True

        except Exception as exc:
            logger.error(f"Excel 리포트 생성 실패: {exc}", exc_info=True)
            return False

    def _generate_summary_sheet(self, wb: Workbook, result: ImpactAnalysisResult):
        """Excel 요약 시트 생성"""
        ws = wb.create_sheet("Summary")

        # 스타일 정의
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 제목
        ws["A1"] = "영향도 분석 요약"
        ws["A1"].font = title_font

        # 분석 정보
        row = 3
        ws[f"A{row}"] = "분석 대상"
        ws[f"B{row}"] = result.target_name
        ws[f"A{row}"].font = header_font

        row += 1
        ws[f"A{row}"] = "프로젝트"
        ws[f"B{row}"] = result.project_name or "전체"

        row += 1
        ws[f"A{row}"] = "분석 유형"
        ws[f"B{row}"] = result.analysis_type

        row += 1
        ws[f"A{row}"] = "분석 일시"
        ws[f"B{row}"] = result.timestamp

        # 영향 범위
        row += 2
        ws[f"A{row}"] = "총 영향 범위"
        ws[f"A{row}"].font = title_font

        row += 1
        ws[f"A{row}"] = "영향받는 클래스"
        ws[f"B{row}"] = result.summary.total_impacted_classes

        row += 1
        ws[f"A{row}"] = "영향받는 메서드"
        ws[f"B{row}"] = result.summary.total_impacted_methods

        row += 1
        ws[f"A{row}"] = "영향받는 패키지"
        ws[f"B{row}"] = result.summary.total_impacted_packages

        row += 1
        ws[f"A{row}"] = "최대 호출 깊이"
        ws[f"B{row}"] = result.summary.max_depth

        row += 1
        ws[f"A{row}"] = "평균 호출 깊이"
        ws[f"B{row}"] = result.summary.avg_depth

        # 리스크 분포
        row += 2
        ws[f"A{row}"] = "리스크 분포"
        ws[f"A{row}"].font = title_font

        row += 1
        ws[f"A{row}"] = "High"
        ws[f"B{row}"] = result.summary.risk_distribution.get("HIGH", 0)
        ws[f"B{row}"].font = Font(color="FF0000", bold=True)

        row += 1
        ws[f"A{row}"] = "Medium"
        ws[f"B{row}"] = result.summary.risk_distribution.get("MEDIUM", 0)
        ws[f"B{row}"].font = Font(color="FFA500", bold=True)

        row += 1
        ws[f"A{row}"] = "Low"
        ws[f"B{row}"] = result.summary.risk_distribution.get("LOW", 0)
        ws[f"B{row}"].font = Font(color="008000", bold=True)

        # 순환 참조
        if result.has_circular_reference:
            row += 2
            ws[f"A{row}"] = "순환 참조 감지"
            ws[f"B{row}"] = f"{len(result.circular_paths)}개"
            ws[f"B{row}"].font = Font(color="FF0000", bold=True)

        # 열 너비 조정
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _generate_detail_sheet(self, wb: Workbook, result: ImpactAnalysisResult):
        """Excel 상세 시트 생성"""
        ws = wb.create_sheet("Impact Detail")

        # 헤더
        headers = ["Level", "Depth", "Package", "Class", "Method", "SQL Type", "Complexity", "Risk", "비고"]
        ws.append(headers)

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터 추가
        row_num = 2
        for level in sorted(result.impact_tree.keys()):
            nodes = result.impact_tree[level]
            for node in nodes:
                remark = "직접 사용" if level == 1 else f"{level-1}차 호출자"

                ws.cell(row_num, 1, level)
                ws.cell(row_num, 2, node.depth)
                ws.cell(row_num, 3, node.package_name)
                ws.cell(row_num, 4, node.class_name)
                ws.cell(row_num, 5, node.method_name)
                ws.cell(row_num, 6, node.sql_type or "")
                ws.cell(row_num, 7, node.sql_complexity or "")
                ws.cell(row_num, 8, node.risk_grade)
                ws.cell(row_num, 9, remark)

                # Risk 셀 색상
                risk_cell = ws.cell(row_num, 8)
                if node.risk_grade == "HIGH":
                    risk_cell.font = Font(color="FF0000", bold=True)
                elif node.risk_grade == "MEDIUM":
                    risk_cell.font = Font(color="FFA500", bold=True)
                else:
                    risk_cell.font = Font(color="008000")

                row_num += 1

        # 열 너비 조정
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 15

    def _generate_package_sheet(self, wb: Workbook, result: ImpactAnalysisResult):
        """Excel 패키지 시트 생성"""
        ws = wb.create_sheet("Package Summary")

        # 헤더
        headers = ["Package", "영향 클래스 수", "영향 메서드 수", "평균 Depth", "High", "Medium", "Low"]
        ws.append(headers)

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터 추가
        for row_num, pkg in enumerate(result.package_summary, 2):
            ws.cell(row_num, 1, pkg.package_name)
            ws.cell(row_num, 2, pkg.impacted_classes)
            ws.cell(row_num, 3, pkg.impacted_methods)
            ws.cell(row_num, 4, pkg.avg_depth)
            ws.cell(row_num, 5, pkg.risk_distribution.get("HIGH", 0))
            ws.cell(row_num, 6, pkg.risk_distribution.get("MEDIUM", 0))
            ws.cell(row_num, 7, pkg.risk_distribution.get("LOW", 0))

        # 열 너비 조정
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

    def _generate_sql_sheet(self, wb: Workbook, result: ImpactAnalysisResult):
        """Excel SQL 상세 시트 생성"""
        ws = wb.create_sheet("SQL Detail")

        # 헤더
        headers = ["SQL ID", "SQL Type", "Mapper Class", "Mapper Method", "Complexity", "Query Preview"]
        ws.append(headers)

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터 추가
        for row_num, sql in enumerate(result.sql_details, 2):
            ws.cell(row_num, 1, sql.sql_id)
            ws.cell(row_num, 2, sql.sql_type)
            ws.cell(row_num, 3, sql.mapper_class)
            ws.cell(row_num, 4, sql.mapper_method)
            ws.cell(row_num, 5, sql.complexity)
            ws.cell(row_num, 6, sql.query_preview or "")

        # 열 너비 조정
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 50

    def _generate_test_scope_sheet(self, wb: Workbook, result: ImpactAnalysisResult):
        """Excel 테스트 범위 시트 생성"""
        ws = wb.create_sheet("Test Scope")

        # 헤더
        headers = ["영향받는 클래스", "대응 테스트 클래스", "테스트 메서드 수", "상태"]
        ws.append(headers)

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터 추가
        for row_num, item in enumerate(result.test_scope, 2):
            ws.cell(row_num, 1, item.impacted_class)
            ws.cell(row_num, 2, item.test_class or "N/A")
            ws.cell(row_num, 3, item.test_method_count)
            ws.cell(row_num, 4, item.status)

            # 상태 셀 색상
            status_cell = ws.cell(row_num, 4)
            if item.status == "존재":
                status_cell.font = Font(color="008000", bold=True)
            else:
                status_cell.font = Font(color="FF0000", bold=True)

        # 열 너비 조정
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12

    def generate_json(
        self,
        result: ImpactAnalysisResult,
        filepath: Path,
    ) -> bool:
        """JSON 리포트 생성

        Args:
            result: ImpactAnalysisResult 객체
            filepath: 저장 경로

        Returns:
            성공 여부
        """
        try:
            # Pydantic 모델을 JSON으로 직렬화
            data = result.model_dump(mode="json")

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            logger.info(f"JSON 리포트 생성 완료: {filepath}")
            return True

        except Exception as exc:
            logger.error(f"JSON 리포트 생성 실패: {exc}", exc_info=True)
            return False

    def generate_mermaid_diagram(
        self,
        result: ImpactAnalysisResult,
        filepath: Path,
    ) -> bool:
        """Mermaid 다이어그램 생성

        Args:
            result: ImpactAnalysisResult 객체
            filepath: 저장 경로 (.md)

        Returns:
            성공 여부
        """
        try:
            generator = ImpactMermaidGenerator()
            return generator.generate_diagram(result, filepath)

        except Exception as exc:
            logger.error(f"Mermaid 다이어그램 생성 실패: {exc}", exc_info=True)
            return False
