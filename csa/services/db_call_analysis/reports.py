from __future__ import annotations

from typing import Optional


class ReportMixin:
    """Generate external reports (Excel, etc.)."""

    def generate_crud_excel(self, project_name: Optional[str] = None, output_file: Optional[str] = None) -> bool:
        """
        Export the CRUD matrix to an Excel workbook.
        """
        try:
            import pandas as pd  # type: ignore
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore

            crud_data = self.generate_crud_table_matrix(project_name)
            if "error" in crud_data:
                self.logger.error(f"Cannot build CRUD matrix: {crud_data['error']}")
                return False

            table_matrix = crud_data["table_matrix"]
            table_names = crud_data["table_names"]
            summary = crud_data["summary"]

            if not table_matrix or not table_names:
                self.logger.warning("No CRUD matrix data available for export.")
                return False

            df_data = []
            for row in table_matrix:
                class_name = row["class_name"]
                package_name = row.get("package_name", "N/A")
                row_data = {"Package": package_name, "Class": class_name}
                for table_name in table_names:
                    row_data[table_name] = row.get(table_name, "-")
                df_data.append(row_data)

            df = pd.DataFrame(df_data)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "CRUD Matrix"

            for row_idx, row in enumerate(df.itertuples(index=False), start=3):
                worksheet.cell(row=row_idx, column=1, value=row.Package)
                worksheet.cell(row=row_idx, column=2, value=row.Class)
                for col_idx, table_name in enumerate(table_names, start=3):
                    value = getattr(row, table_name, "-")
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            worksheet.cell(row=1, column=1, value="Package")
            worksheet.cell(row=1, column=2, value="Class")
            for idx, table_name in enumerate(table_names, start=3):
                worksheet.cell(row=1, column=idx, value=f"Table: {table_name}")
                worksheet.cell(row=2, column=idx, value="CRUD")
            worksheet.cell(row=2, column=1, value="Package")
            worksheet.cell(row=2, column=2, value="Class")

            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            class_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col in range(1, len(table_names) + 3):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for col in range(1, len(table_names) + 3):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for row in range(3, len(table_matrix) + 3):
                for col in range(1, len(table_names) + 3):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if col <= 2:
                        cell.fill = class_fill
                        cell.font = Font(bold=False)
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment
                        cell.font = Font(bold=bool(cell.value and cell.value != "-"))

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except Exception:  # pylint: disable=broad-except
                        continue
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 20)

            summary_sheet = workbook.create_sheet("Summary")
            summary_data = [
                ["CRUD Matrix Summary", ""],
                ["Project Name", project_name],
                ["Total Classes", summary.get("total_classes", 0)],
                ["Total Tables", summary.get("total_tables", 0)],
                ["", ""],
                ["CRUD Operations", ""],
                ["Create (C)", summary.get("crud_stats", {}).get("C", 0)],
                ["Read (R)", summary.get("crud_stats", {}).get("R", 0)],
                ["Update (U)", summary.get("crud_stats", {}).get("U", 0)],
                ["Delete (D)", summary.get("crud_stats", {}).get("D", 0)],
                ["Other (O)", summary.get("crud_stats", {}).get("O", 0)],
                ["", ""],
                ["Most Active Class", summary.get("most_active_class", "N/A")],
                ["Most Used Table", summary.get("most_used_table", "N/A")],
            ]

            for row in summary_data:
                summary_sheet.append(row)

            for row in range(1, len(summary_data) + 1):
                for col in range(1, 3):
                    cell = summary_sheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if row == 1:
                        cell.font = Font(bold=True, size=14)
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    elif col == 1:
                        cell.font = Font(bold=True)
                        cell.fill = class_fill

            summary_sheet.column_dimensions["A"].width = 20
            summary_sheet.column_dimensions["B"].width = 15

            workbook.save(output_file)
            self.logger.info(f"Excel report created: {output_file}")
            return True
        except ImportError as exc:
            self.logger.error(f"Required libraries not installed: {exc}")
            self.logger.error("Install them with: pip install pandas openpyxl")
            return False
        except Exception as exc:  # pylint: disable=broad-except
            self.logger.error(f"Excel report generation error: {exc}")
            return False
