"""
데이터베이스 호출관계 분석 서비스

이 모듈은 어플리케이션과 데이터베이스 간의 호출관계를 분석하고 시각화하는 기능을 제공합니다.
- Controller → Service → Repository → SQL → Table/Column 호출 체인 분석
- CRUD 매트릭스 생성
- 호출 관계 그래프 생성 (Mermaid 다이어그램)
- 존재하지 않는 Table/Column 노드 식별 및 시각화
- 영향도 분석
"""

import json
from typing import List, Dict, Any, Optional, Set, Tuple
from neo4j import Driver
from src.utils.logger import get_logger


class DBCallAnalysisService:
    """데이터베이스 호출관계 분석 서비스"""
    
    def __init__(self, driver: Driver):
        """Neo4j 드라이버로 초기화"""
        self.driver = driver
        self.logger = get_logger(__name__)
    
    def analyze_call_chain(self, project_name: str = None, start_class: str = None, start_method: str = None) -> Dict[str, Any]:
        """
        Controller → Service → Repository → SQL → Table/Column 호출 체인을 분석합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            start_class: 시작 클래스 (선택사항)
            start_method: 시작 메서드 (선택사항)
            
        Returns:
            호출 체인 분석 결과
        """
        try:
            with self.driver.session() as session:
                if start_class and start_method:
                    # 특정 메서드부터 시작하는 호출 체인
                    call_chain = self._get_method_call_chain(session, start_class, start_method)
                elif start_class:
                    # 특정 클래스부터 시작하는 호출 체인
                    call_chain = self._get_class_call_chain(session, start_class)
                else:
                    # 전체 프로젝트의 호출 체인
                    call_chain = self._get_project_call_chain(session)
                
                # 존재하지 않는 Table/Column 노드 식별
                missing_nodes = self._identify_missing_nodes(session, call_chain)
                
                return {
                    'project_name': project_name,
                    'call_chain': call_chain,
                    'missing_nodes': missing_nodes,
                    'analysis_summary': self._generate_analysis_summary(call_chain, missing_nodes)
                }
                
        except Exception as e:
            self.logger.error(f"호출 체인 분석 오류: {str(e)}")
            return {'error': str(e)}
    
    def _get_method_call_chain(self, session, class_name: str, method_name: str) -> List[Dict[str, Any]]:
        """특정 메서드부터 시작하는 호출 체인을 분석합니다."""
        query = """
        MATCH (c:Class {name: $class_name})-[:HAS_METHOD]->(m:Method {name: $method_name})
        OPTIONAL MATCH (m)-[:CALLS*0..5]->(target_method:Method)
        OPTIONAL MATCH (target_method)<-[:HAS_METHOD]-(target_class:Class)
        OPTIONAL MATCH (target_method)-[:CALLS]->(sql:SqlStatement)
        OPTIONAL MATCH (sql)-[:USES_TABLE]->(table:Table)
        OPTIONAL MATCH (sql)-[:USES_COLUMN]->(column:Column)
        RETURN m.name as source_method,
               c.name as source_class,
               c.package_name as source_package,
               target_method.name as target_method,
               target_class.name as target_class,
               target_class.package_name as target_package,
               sql.id as sql_id,
               sql.sql_type as sql_type,
               sql.tables as sql_tables,
               sql.columns as sql_columns,
               table.name as table_name,
               column.name as column_name,
               column.table_name as column_table_name
        ORDER BY source_method, target_class.name, target_method.name
        """
        
        result = session.run(query, 
                           class_name=class_name, 
                           method_name=method_name)
        
        call_chain = []
        for record in result:
            call_chain.append({
                'source_method': record['source_method'],
                'source_class': record['source_class'],
                'source_package': record['source_package'] or 'default',
                'target_method': record['target_method'],
                'target_class': record['target_class'],
                'target_package': record['target_package'] or 'default',
                'sql_id': record['sql_id'],
                'sql_type': record['sql_type'],
                'sql_tables': json.loads(record['sql_tables']) if record['sql_tables'] else [],
                'sql_columns': json.loads(record['sql_columns']) if record['sql_columns'] else [],
                'table_name': record['table_name'],
                'column_name': record['column_name'],
                'column_table_name': record['column_table_name']
            })
        
        return call_chain
    
    def _get_class_call_chain(self, session, class_name: str) -> List[Dict[str, Any]]:
        """특정 클래스부터 시작하는 호출 체인을 분석합니다."""
        query = """
        MATCH (c:Class {name: $class_name})-[:HAS_METHOD]->(m:Method)
        OPTIONAL MATCH (m)-[:CALLS*0..5]->(target_method:Method)
        OPTIONAL MATCH (target_method)<-[:HAS_METHOD]-(target_class:Class)
        OPTIONAL MATCH (target_method)-[:CALLS]->(sql:SqlStatement)
        OPTIONAL MATCH (sql)-[:USES_TABLE]->(table:Table)
        OPTIONAL MATCH (sql)-[:USES_COLUMN]->(column:Column)
        RETURN m.name as source_method,
               c.name as source_class,
               c.package_name as source_package,
               target_method.name as target_method,
               target_class.name as target_class,
               target_class.package_name as target_package,
               sql.id as sql_id,
               sql.sql_type as sql_type,
               sql.tables as sql_tables,
               sql.columns as sql_columns,
               table.name as table_name,
               column.name as column_name,
               column.table_name as column_table_name
        ORDER BY source_method, target_class.name, target_method.name
        """
        
        result = session.run(query, 
                           class_name=class_name)
        
        call_chain = []
        for record in result:
            call_chain.append({
                'source_method': record['source_method'],
                'source_class': record['source_class'],
                'source_package': record['source_package'] or 'default',
                'target_method': record['target_method'],
                'target_class': record['target_class'],
                'target_package': record['target_package'] or 'default',
                'sql_id': record['sql_id'],
                'sql_type': record['sql_type'],
                'sql_tables': json.loads(record['sql_tables']) if record['sql_tables'] else [],
                'sql_columns': json.loads(record['sql_columns']) if record['sql_columns'] else [],
                'table_name': record['table_name'],
                'column_name': record['column_name'],
                'column_table_name': record['column_table_name']
            })
        
        return call_chain
    
    def _get_project_call_chain(self, session) -> List[Dict[str, Any]]:
        """전체 프로젝트의 호출 체인을 분석합니다."""
        query = """
        MATCH (c:Class)-[:HAS_METHOD]->(m:Method)
        OPTIONAL MATCH (m)-[:CALLS*0..5]->(target_method:Method)
        OPTIONAL MATCH (target_method)<-[:HAS_METHOD]-(target_class:Class)
        OPTIONAL MATCH (target_method)-[:CALLS]->(sql:SqlStatement)
        OPTIONAL MATCH (sql)-[:USES_TABLE]->(table:Table)
        OPTIONAL MATCH (sql)-[:USES_COLUMN]->(column:Column)
        RETURN m.name as source_method,
               c.name as source_class,
               c.package_name as source_package,
               target_method.name as target_method,
               target_class.name as target_class,
               target_class.package_name as target_package,
               sql.id as sql_id,
               sql.sql_type as sql_type,
               sql.tables as sql_tables,
               sql.columns as sql_columns,
               table.name as table_name,
               column.name as column_name,
               column.table_name as column_table_name
        ORDER BY source_class, source_method, target_class.name, target_method.name
        """
        
        result = session.run(query)
        
        call_chain = []
        for record in result:
            call_chain.append({
                'source_method': record['source_method'],
                'source_class': record['source_class'],
                'source_package': record['source_package'] or 'default',
                'target_method': record['target_method'],
                'target_class': record['target_class'],
                'target_package': record['target_package'] or 'default',
                'sql_id': record['sql_id'],
                'sql_type': record['sql_type'],
                'sql_tables': json.loads(record['sql_tables']) if record['sql_tables'] else [],
                'sql_columns': json.loads(record['sql_columns']) if record['sql_columns'] else [],
                'table_name': record['table_name'],
                'column_name': record['column_name'],
                'column_table_name': record['column_table_name']
            })
        
        return call_chain
    
    def _identify_missing_nodes(self, session, call_chain: List[Dict[str, Any]]) -> Dict[str, List[str]]:
        """존재하지 않는 Table/Column 노드를 식별합니다."""
        missing_tables = set()
        missing_columns = set()
        
        # SQL에서 참조하는 테이블과 컬럼 수집
        referenced_tables = set()
        referenced_columns = set()
        
        for call in call_chain:
            # SQL 테이블 정보 수집
            if call.get('sql_tables'):
                for table_info in call['sql_tables']:
                    if isinstance(table_info, dict) and 'name' in table_info:
                        referenced_tables.add(table_info['name'])
            
            # SQL 컬럼 정보 수집
            if call.get('sql_columns'):
                for column_info in call['sql_columns']:
                    if isinstance(column_info, dict) and 'name' in column_info:
                        referenced_columns.add(column_info['name'])
                        if 'table' in column_info and column_info['table']:
                            referenced_tables.add(column_info['table'])
        
        # 실제 존재하는 테이블과 컬럼 조회
        existing_tables = self._get_existing_tables(session)
        existing_columns = self._get_existing_columns(session)
        
        # 존재하지 않는 테이블 식별
        for table_name in referenced_tables:
            if table_name not in existing_tables:
                missing_tables.add(table_name)
        
        # 존재하지 않는 컬럼 식별
        for column_name in referenced_columns:
            if column_name not in existing_columns:
                missing_columns.add(column_name)
        
        return {
            'missing_tables': list(missing_tables),
            'missing_columns': list(missing_columns)
        }
    
    def _get_existing_tables(self, session) -> Set[str]:
        """실제 존재하는 테이블 목록을 조회합니다."""
        query = """
        MATCH (t:Table)
        RETURN t.name as table_name
        """
        
        result = session.run(query)
        return {record['table_name'] for record in result}
    
    def _get_existing_columns(self, session) -> Set[str]:
        """실제 존재하는 컬럼 목록을 조회합니다."""
        query = """
        MATCH (c:Column)
        RETURN c.name as column_name
        """
        
        result = session.run(query)
        return {record['column_name'] for record in result}
    
    def _generate_analysis_summary(self, call_chain: List[Dict[str, Any]], missing_nodes: Dict[str, List[str]]) -> Dict[str, Any]:
        """분석 결과 요약을 생성합니다."""
        # 클래스별 통계
        class_stats = {}
        for call in call_chain:
            source_class = call['source_class']
            if source_class not in class_stats:
                class_stats[source_class] = {
                    'method_count': 0,
                    'sql_count': 0,
                    'table_count': 0,
                    'column_count': 0
                }
            
            if call['source_method']:
                class_stats[source_class]['method_count'] += 1
            if call['sql_id']:
                class_stats[source_class]['sql_count'] += 1
            if call['table_name']:
                class_stats[source_class]['table_count'] += 1
            if call['column_name']:
                class_stats[source_class]['column_count'] += 1
        
        return {
            'total_calls': len(call_chain),
            'unique_classes': len(class_stats),
            'unique_methods': len(set(call['source_method'] for call in call_chain if call['source_method'])),
            'unique_sql_statements': len(set(call['sql_id'] for call in call_chain if call['sql_id'])),
            'unique_tables': len(set(call['table_name'] for call in call_chain if call['table_name'])),
            'unique_columns': len(set(call['column_name'] for call in call_chain if call['column_name'])),
            'missing_tables_count': len(missing_nodes['missing_tables']),
            'missing_columns_count': len(missing_nodes['missing_columns']),
            'class_stats': class_stats
        }
    
    def generate_crud_matrix(self, project_name: str = None) -> Dict[str, Any]:
        """
        CRUD 매트릭스를 생성합니다.
        SQL을 직접 호출하는 클래스만 포함합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            
        Returns:
            CRUD 매트릭스 데이터
        """
        try:
            with self.driver.session() as session:
                # SQL을 직접 호출하는 클래스와 실제 테이블 정보를 가져오는 쿼리
                class_crud_query = """
                MATCH (c:Class)-[:HAS_METHOD]->(m:Method)
                MATCH (m)-[:CALLS]->(sql:SqlStatement)
                WHERE sql.tables IS NOT NULL AND sql.tables <> '[]'
                WITH c, m, sql,
                     CASE 
                       WHEN sql.sql_type = 'SELECT' THEN 'R'
                       WHEN sql.sql_type = 'INSERT' THEN 'C'
                       WHEN sql.sql_type = 'UPDATE' THEN 'U'
                       WHEN sql.sql_type = 'DELETE' THEN 'D'
                       ELSE 'O'
                     END as crud_operation
                RETURN c.name as class_name,
                       c.package_name as package_name,
                       sql.tables as tables_json,
                       crud_operation as operation,
                       sql.id as sql_id
                ORDER BY c.name
                """
                
                result = session.run(class_crud_query)
                raw_data = [record.data() for record in result]
                
                # 클래스별로 그룹화하여 매트릭스 생성 (1:1 관계로 표시)
                class_table_relations = []
                processed_combinations = set()
                
                for row in raw_data:
                    class_name = row['class_name']
                    package_name = row['package_name']
                    operation = row['operation']
                    sql_id = row['sql_id']
                    database_name = 'default'  # 기본값 설정
                    schema_name = 'public'    # 기본값 설정
                    
                    # 실제 테이블 정보 파싱
                    try:
                        tables_json = row['tables_json']
                        if tables_json and tables_json != '[]':
                            tables = json.loads(tables_json)
                            for table_info in tables:
                                if isinstance(table_info, dict) and 'name' in table_info:
                                    table_name = table_info['name']
                                    
                                    # 클래스-테이블 조합이 이미 처리되었는지 확인
                                    combination_key = f"{class_name}_{table_name}"
                                    if combination_key not in processed_combinations:
                                        processed_combinations.add(combination_key)
                                        
                                        # 해당 클래스-테이블 조합의 모든 SQL 찾기
                                        table_operations = set()
                                        table_sql_statements = set()
                                        
                                        for check_row in raw_data:
                                            if (check_row['class_name'] == class_name and 
                                                check_row['tables_json'] and 
                                                check_row['tables_json'] != '[]'):
                                                try:
                                                    check_tables = json.loads(check_row['tables_json'])
                                                    for check_table in check_tables:
                                                        if (isinstance(check_table, dict) and 
                                                            'name' in check_table and 
                                                            check_table['name'] == table_name):
                                                            table_operations.add(check_row['operation'])
                                                            table_sql_statements.add(check_row['sql_id'])
                                                except (json.JSONDecodeError, TypeError):
                                                    continue
                                        
                                        class_table_relations.append({
                                            'class_name': class_name,
                                            'package_name': package_name,
                                            'table_name': table_name,
                                            'database_name': database_name,
                                            'schema_name': schema_name,
                                            'operations': list(table_operations),
                                            'sql_statements': list(table_sql_statements)
                                        })
                    except (json.JSONDecodeError, TypeError) as e:
                        self.logger.warning(f"테이블 JSON 파싱 오류: {e}")
                        continue
                
                # 클래스-테이블 관계를 클래스별로 그룹화
                class_matrix = {}
                for relation in class_table_relations:
                    class_name = relation['class_name']
                    if class_name not in class_matrix:
                        class_matrix[class_name] = {
                            'class_name': class_name,
                            'package_name': relation['package_name'],
                            'tables': [],
                            'operations': set(),
                            'sql_statements': set()
                        }
                    
                    class_matrix[class_name]['tables'].append({
                        'table_name': relation['table_name'],
                        'database_name': relation['database_name'],
                        'schema_name': relation['schema_name'],
                        'operations': relation['operations']
                    })
                    if isinstance(relation['operations'], dict):
                        class_matrix[class_name]['operations'].update(relation['operations'])
                    elif isinstance(relation['operations'], (list, set)):
                        class_matrix[class_name]['operations'].update(relation['operations'])
                    
                    if isinstance(relation['sql_statements'], dict):
                        class_matrix[class_name]['sql_statements'].update(relation['sql_statements'])
                    elif isinstance(relation['sql_statements'], (list, set)):
                        class_matrix[class_name]['sql_statements'].update(relation['sql_statements'])
                
                # 최종 형태로 변환
                class_matrix = [
                    {
                        'class_name': data['class_name'],
                        'package_name': data['package_name'],
                        'tables': data['tables'],
                        'operations': list(data['operations']),
                        'sql_statements': list(data['sql_statements'])
                    }
                    for data in class_matrix.values()
                ]
                
                # 테이블별 CRUD 매트릭스 (Python에서 처리)
                table_crud_query = """
                MATCH (sql:SqlStatement)
                WHERE sql.tables IS NOT NULL AND sql.tables <> '[]'
                RETURN sql.tables as tables_json, sql.sql_type as operation
                """
                
                result = session.run(table_crud_query)
                raw_table_data = [record.data() for record in result]
                
                # Python에서 테이블별 CRUD 매트릭스 생성
                table_stats = {}
                for row in raw_table_data:
                    try:
                        tables_json = row['tables_json']
                        operation = row['operation']
                        
                        if tables_json and tables_json != '[]':
                            tables = json.loads(tables_json)
                            for table_info in tables:
                                if isinstance(table_info, dict) and 'name' in table_info:
                                    table_name = table_info['name']
                                    if table_name not in table_stats:
                                        table_stats[table_name] = {}
                                    if operation not in table_stats[table_name]:
                                        table_stats[table_name][operation] = 0
                                    table_stats[table_name][operation] += 1
                    except (json.JSONDecodeError, TypeError) as e:
                        self.logger.warning(f"테이블 JSON 파싱 오류: {e}")
                        continue
                
                # 테이블별 CRUD 매트릭스 형식으로 변환
                table_matrix = []
                for table_name, operations in table_stats.items():
                    operations_list = [{'operation': op, 'count': count} for op, count in operations.items()]
                    table_matrix.append({
                        'table_name': table_name,
                        'operations': operations_list
                    })
                
                return {
                    'project_name': project_name,
                    'class_matrix': class_matrix,
                    'table_matrix': table_matrix,
                    'summary': self._generate_crud_summary(class_matrix, table_matrix)
                }
                
        except Exception as e:
            self.logger.error(f"CRUD 매트릭스 생성 오류: {str(e)}")
            return {'error': str(e)}
    
    def _generate_crud_summary(self, class_matrix: List[Dict], table_matrix: List[Dict]) -> Dict[str, Any]:
        """CRUD 매트릭스 요약을 생성합니다."""
        total_classes = len(class_matrix)
        total_tables = len(table_matrix)
        
        # CRUD 작업별 통계
        crud_stats = {'C': 0, 'R': 0, 'U': 0, 'D': 0, 'O': 0}
        for class_data in class_matrix:
            for operation in class_data['operations']:
                if operation in crud_stats:
                    crud_stats[operation] += 1
        
        return {
            'total_classes': total_classes,
            'total_tables': total_tables,
            'crud_stats': crud_stats,
            'most_active_class': max(class_matrix, key=lambda x: len(x['operations']))['class_name'] if class_matrix else None,
            'most_used_table': max(table_matrix, key=lambda x: sum(op['count'] for op in x['operations']))['table_name'] if table_matrix else None
        }
    
    def generate_call_chain_diagram(self, project_name: str = None, start_class: str = None, start_method: str = None) -> str:
        """
        호출 체인을 Mermaid 다이어그램으로 생성합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            start_class: 시작 클래스 (선택사항)
            start_method: 시작 메서드 (선택사항)
            
        Returns:
            Mermaid 다이어그램 문자열
        """
        try:
            analysis_result = self.analyze_call_chain(project_name, start_class, start_method)
            
            if 'error' in analysis_result:
                return f"오류: {analysis_result['error']}"
            
            call_chain = analysis_result['call_chain']
            missing_nodes = analysis_result['missing_nodes']
            
            if not call_chain:
                return "```mermaid\ngraph TD\n    A[No call chain found]\n```"
            
            # Mermaid 다이어그램 생성
            diagram_lines = ["```mermaid", "graph TD"]
            
            # 노드 정의
            nodes = set()
            for call in call_chain:
                if call['source_class']:
                    nodes.add(call['source_class'])
                if call['target_class']:
                    nodes.add(call['target_class'])
                if call['table_name']:
                    nodes.add(f"Table_{call['table_name']}")
                if call['column_name']:
                    nodes.add(f"Column_{call['column_name']}")
            
            # 존재하지 않는 노드 추가 (적색 점선으로 표시)
            for missing_table in missing_nodes['missing_tables']:
                nodes.add(f"MissingTable_{missing_table}")
            
            for missing_column in missing_nodes['missing_columns']:
                nodes.add(f"MissingColumn_{missing_column}")
            
            # 노드 스타일 정의
            for node in sorted(nodes):
                if node.startswith('MissingTable_'):
                    table_name = node.replace('MissingTable_', '')
                    diagram_lines.append(f"    {node}[\"❌ {table_name}\"]:::missingTable")
                elif node.startswith('MissingColumn_'):
                    column_name = node.replace('MissingColumn_', '')
                    diagram_lines.append(f"    {node}[\"❌ {column_name}\"]:::missingColumn")
                elif node.startswith('Table_'):
                    table_name = node.replace('Table_', '')
                    diagram_lines.append(f"    {node}[\"📊 {table_name}\"]:::table")
                elif node.startswith('Column_'):
                    column_name = node.replace('Column_', '')
                    diagram_lines.append(f"    {node}[\"📋 {column_name}\"]:::column")
                else:
                    diagram_lines.append(f"    {node}[\"🏢 {node}\"]:::class")
            
            # 연결선 정의
            for call in call_chain:
                if call['source_class'] and call['target_class']:
                    diagram_lines.append(f"    {call['source_class']} --> {call['target_class']}")
                
                if call['target_class'] and call['table_name']:
                    table_node = f"Table_{call['table_name']}"
                    diagram_lines.append(f"    {call['target_class']} --> {table_node}")
                
                if call['table_name'] and call['column_name']:
                    table_node = f"Table_{call['table_name']}"
                    column_node = f"Column_{call['column_name']}"
                    diagram_lines.append(f"    {table_node} --> {column_node}")
            
            # 존재하지 않는 노드 연결 (적색 점선)
            for call in call_chain:
                if call['table_name'] and call['table_name'] in missing_nodes['missing_tables']:
                    missing_table_node = f"MissingTable_{call['table_name']}"
                    if call['target_class']:
                        diagram_lines.append(f"    {call['target_class']} -.-> {missing_table_node}")
                
                if call['column_name'] and call['column_name'] in missing_nodes['missing_columns']:
                    missing_column_node = f"MissingColumn_{call['column_name']}"
                    if call['table_name']:
                        table_node = f"Table_{call['table_name']}"
                        diagram_lines.append(f"    {table_node} -.-> {missing_column_node}")
            
            # 스타일 정의
            diagram_lines.extend([
                "",
                "    classDef class fill:#e1f5fe,stroke:#01579b,stroke-width:2px",
                "    classDef table fill:#f3e5f5,stroke:#4a148c,stroke-width:2px",
                "    classDef column fill:#e8f5e8,stroke:#1b5e20,stroke-width:2px",
                "    classDef missingTable fill:#ffebee,stroke:#c62828,stroke-width:2px,stroke-dasharray: 5 5",
                "    classDef missingColumn fill:#fff3e0,stroke:#ef6c00,stroke-width:2px,stroke-dasharray: 5 5"
            ])
            
            diagram_lines.append("```")
            
            return "\n".join(diagram_lines)
            
        except Exception as e:
            self.logger.error(f"호출 체인 다이어그램 생성 오류: {str(e)}")
            return f"오류: {str(e)}"
    
    def analyze_table_impact(self, project_name: str = None, table_name: str = None) -> Dict[str, Any]:
        """
        특정 테이블 변경 시 영향받는 클래스/메서드를 분석합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            table_name: 분석할 테이블 이름
            
        Returns:
            영향도 분석 결과
        """
        try:
            with self.driver.session() as session:
                # 테이블을 사용하는 클래스/메서드 조회
                impact_query = """
                MATCH (c:Class)-[:HAS_METHOD]->(m:Method)
                MATCH (m)-[:CALLS]->(sql:SqlStatement)
                WHERE sql.tables CONTAINS $table_name OR 
                      ANY(table_info IN sql.tables WHERE table_info.name = $table_name)
                RETURN c.name as class_name,
                       c.package_name as package_name,
                       m.name as method_name,
                       sql.id as sql_id,
                       sql.sql_type as sql_type,
                       sql.complexity_score as complexity_score
                ORDER BY c.name, m.name
                """
                
                result = session.run(impact_query, 
                                   project_name=project_name, 
                                   table_name=table_name)
                
                impacted_classes = []
                for record in result:
                    impacted_classes.append({
                        'class_name': record['class_name'],
                        'package_name': record['package_name'],
                        'method_name': record['method_name'],
                        'sql_id': record['sql_id'],
                        'sql_type': record['sql_type'],
                        'complexity_score': record['complexity_score']
                    })
                
                # 영향도 요약
                summary = {
                    'table_name': table_name,
                    'total_impacted_classes': len(set(c['class_name'] for c in impacted_classes)),
                    'total_impacted_methods': len(set(f"{c['class_name']}.{c['method_name']}" for c in impacted_classes)),
                    'total_sql_statements': len(set(c['sql_id'] for c in impacted_classes if c['sql_id'])),
                    'crud_operations': list(set(c['sql_type'] for c in impacted_classes if c['sql_type'])),
                    'high_complexity_sql': [c for c in impacted_classes if c['complexity_score'] and c['complexity_score'] > 7]
                }
                
                return {
                    'table_name': table_name,
                    'impacted_classes': impacted_classes,
                    'summary': summary
                }
                
        except Exception as e:
            self.logger.error(f"테이블 영향도 분석 오류: {str(e)}")
            return {'error': str(e)}
    
    def get_database_usage_statistics(self, project_name: str = None) -> Dict[str, Any]:
        """
        데이터베이스 사용 통계를 조회합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            
        Returns:
            데이터베이스 사용 통계
        """
        try:
            with self.driver.session() as session:
                # SQL 통계
                sql_stats_query = """
                MATCH (sql:SqlStatement)
                RETURN 
                    count(sql) as total_sql,
                    sum(CASE WHEN sql.sql_type = 'SELECT' THEN 1 ELSE 0 END) as SELECT,
                    sum(CASE WHEN sql.sql_type = 'INSERT' THEN 1 ELSE 0 END) as INSERT,
                    sum(CASE WHEN sql.sql_type = 'UPDATE' THEN 1 ELSE 0 END) as UPDATE,
                    sum(CASE WHEN sql.sql_type = 'DELETE' THEN 1 ELSE 0 END) as DELETE
                """
                
                result = session.run(sql_stats_query, project_name=project_name)
                sql_stats = result.single().data() if result.single() else {}
                
                # 테이블 사용 통계
                table_usage_query = """
                MATCH (sql:SqlStatement {project_name: $project_name})
                WHERE sql.tables IS NOT NULL
                UNWIND sql.tables as table_info
                WITH table_info.name as table_name, sql.sql_type as operation
                RETURN 
                    table_name,
                    count(*) as access_count,
                    collect(DISTINCT operation) as operations
                ORDER BY access_count DESC
                """
                
                result = session.run(table_usage_query, project_name=project_name)
                table_usage = [record.data() for record in result]
                
                # 복잡도 통계
                complexity_query = """
                MATCH (sql:SqlStatement {project_name: $project_name})
                WHERE sql.complexity_score IS NOT NULL
                WITH sql.complexity_score as score,
                     CASE 
                         WHEN sql.complexity_score <= 3 THEN 'simple'
                         WHEN sql.complexity_score <= 7 THEN 'medium'
                         WHEN sql.complexity_score <= 12 THEN 'complex'
                         ELSE 'very_complex'
                     END as complexity_level
                RETURN 
                    complexity_level,
                    count(*) as count
                """
                
                result = session.run(complexity_query, project_name=project_name)
                complexity_stats = {record['complexity_level']: record['count'] for record in result}
                
                return {
                    'project_name': project_name,
                    'sql_statistics': sql_stats,
                    'table_usage': table_usage,
                    'complexity_statistics': complexity_stats
                }
                
        except Exception as e:
            self.logger.error(f"데이터베이스 사용 통계 조회 오류: {str(e)}")
            return {'error': str(e)}
    
    def generate_crud_visualization_diagram(self, project_name: str = None) -> str:
        """
        CRUD 매트릭스를 기반으로 어플리케이션-데이터베이스 호출관계를 Mermaid 다이어그램으로 생성합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            
        Returns:
            Mermaid 다이어그램 문자열
        """
        try:
            # CRUD 매트릭스 데이터 가져오기
            crud_data = self.generate_crud_matrix(project_name)
            class_matrix = crud_data['class_matrix']
            table_matrix = crud_data['table_matrix']
            
            if not class_matrix:
                return "```mermaid\ngraph TD\n    A[No database calls found]\n```"
            
            # Mermaid 다이어그램 생성
            diagram_lines = ["```mermaid", "graph TD"]
            
            # 노드 정의
            nodes = set()
            class_table_relations = []
            
            # 클래스와 테이블 노드 수집
            for class_data in class_matrix:
                class_name = class_data['class_name']
                nodes.add(f"Class_{class_name}")
                
                if 'tables' in class_data and class_data['tables']:
                    for table_info in class_data['tables']:
                        if isinstance(table_info, dict) and 'table_name' in table_info:
                            table_name = table_info['table_name']
                            nodes.add(f"Table_{table_name}")
                            
                            # 클래스-테이블 관계 저장
                            operations = table_info.get('operations', [])
                            class_table_relations.append({
                                'class_name': class_name,
                                'table_name': table_name,
                                'operations': operations,
                                'database_name': table_info.get('database_name', 'default'),
                                'schema_name': table_info.get('schema_name', 'public')
                            })
            
            # 노드 정의 추가
            for node in sorted(nodes):
                if node.startswith('Class_'):
                    class_name = node.replace('Class_', '')
                    diagram_lines.append(f"    {node}[\"🏢 {class_name}\"]")
                elif node.startswith('Table_'):
                    table_name = node.replace('Table_', '')
                    diagram_lines.append(f"    {node}[\"📊 {table_name}\"]")
            
            # 클래스-테이블 관계 연결
            for relation in class_table_relations:
                class_node = f"Class_{relation['class_name']}"
                table_node = f"Table_{relation['table_name']}"
                
                # CRUD 연산을 라벨로 표시
                operations_str = ', '.join(relation['operations'])
                diagram_lines.append(f"    {class_node} -->|{operations_str}| {table_node}")
            
            # 스타일 정의 (Mermaid 호환성을 위해 제거)
            
            diagram_lines.append("```")
            
            return "\n".join(diagram_lines)
            
        except Exception as e:
            self.logger.error(f"CRUD 시각화 다이어그램 생성 오류: {e}")
            return f"```mermaid\ngraph TD\n    A[Error: {str(e)}]\n```"
    
    def generate_crud_table_matrix(self, project_name: str = None) -> Dict[str, Any]:
        """
        CRUD 매트릭스를 표 형태로 생성합니다.
        가로축: 테이블, 세로축: 클래스
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            
        Returns:
            표 형태의 CRUD 매트릭스 데이터
        """
        try:
            # CRUD 매트릭스 데이터 가져오기
            crud_data = self.generate_crud_matrix(project_name)
            class_matrix = crud_data['class_matrix']
            table_matrix = crud_data['table_matrix']
            
            if not class_matrix:
                return {
                    'table_matrix': [],
                    'class_names': [],
                    'table_names': [],
                    'summary': {'total_classes': 0, 'total_tables': 0}
                }
            
            # 모든 테이블 이름 수집
            all_tables = set()
            for class_data in class_matrix:
                if 'tables' in class_data and class_data['tables']:
                    for table_info in class_data['tables']:
                        if isinstance(table_info, dict) and 'table_name' in table_info:
                            all_tables.add(table_info['table_name'])
            
            table_names = sorted(list(all_tables))
            class_names = [class_data['class_name'] for class_data in class_matrix]
            
            # 표 매트릭스 생성
            table_matrix = []
            for class_data in class_matrix:
                class_name = class_data['class_name']
                package_name = class_data.get('package_name', 'N/A')
                row = {
                    'class_name': class_name,
                    'package_name': package_name
                }
                
                # 각 테이블에 대한 CRUD 연산 확인
                class_tables = {}
                if 'tables' in class_data and class_data['tables']:
                    for table_info in class_data['tables']:
                        if isinstance(table_info, dict) and 'table_name' in table_info:
                            table_name = table_info['table_name']
                            operations = table_info.get('operations', [])
                            schema_name = table_info.get('schema_name', 'public')
                            database_name = table_info.get('database_name', 'default')
                            class_tables[table_name] = {
                                'operations': operations,
                                'schema_name': schema_name,
                                'database_name': database_name
                            }
                
                # 모든 테이블에 대해 CRUD 연산 표시
                for table_name in table_names:
                    if table_name in class_tables:
                        table_info = class_tables[table_name]
                        operations = table_info['operations']
                        # CRUD 연산을 정렬하여 표시 (스키마명 제거)
                        sorted_ops = sorted(operations)
                        operations_str = ', '.join(sorted_ops) if sorted_ops else '-'
                        row[table_name] = operations_str
                    else:
                        row[table_name] = '-'
                
                table_matrix.append(row)
            
            return {
                'table_matrix': table_matrix,
                'class_names': class_names,
                'table_names': table_names,
                'summary': crud_data['summary']
            }
            
        except Exception as e:
            self.logger.error(f"CRUD 표 매트릭스 생성 오류: {e}")
            return {'error': str(e)}
    
    def generate_crud_excel(self, project_name: str = None, output_file: str = None) -> bool:
        """
        CRUD 매트릭스를 Excel 파일로 생성합니다.
        
        Args:
            project_name: 프로젝트 이름 (선택사항, 크로스 프로젝트 분석을 위해)
            output_file: 출력할 Excel 파일 경로
            
        Returns:
            성공 여부
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # CRUD 매트릭스 데이터 가져오기
            crud_data = self.generate_crud_table_matrix(project_name)
            
            if 'error' in crud_data:
                self.logger.error(f"CRUD 데이터 가져오기 오류: {crud_data['error']}")
                return False
            
            table_matrix = crud_data['table_matrix']
            class_names = crud_data['class_names']
            table_names = crud_data['table_names']
            summary = crud_data['summary']
            
            if not table_matrix or not table_names:
                self.logger.warning("CRUD 매트릭스 데이터가 없습니다.")
                return False
            
            # DataFrame 생성 (Package를 첫 번째 컬럼으로)
            df_data = []
            for row in table_matrix:
                class_name = row['class_name']
                package_name = row.get('package_name', 'N/A')
                row_data = {
                    'Package': package_name,
                    'Class': class_name
                }
                for table_name in table_names:
                    operations = row.get(table_name, '-')
                    row_data[table_name] = operations
                df_data.append(row_data)
            
            df = pd.DataFrame(df_data)
            
            # Excel 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "CRUD Matrix"
            
            # 스키마 정보 수집 (원본 데이터에서 직접 추출)
            schema_info = {}
            # 원본 CRUD 데이터에서 스키마 정보 추출
            original_crud_data = self.generate_crud_matrix(project_name)
            original_class_matrix = original_crud_data['class_matrix']
            
            for class_data in original_class_matrix:
                if 'tables' in class_data and class_data['tables']:
                    for table_info in class_data['tables']:
                        if isinstance(table_info, dict) and 'table_name' in table_info:
                            table_name = table_info['table_name']
                            schema_name = table_info.get('schema_name', 'public')
                            schema_info[table_name] = schema_name
            
            # 첫 번째 행: 스키마 정보
            schema_row = ['', '']  # Package, Class 컬럼은 비움
            for table_name in table_names:
                schema = schema_info.get(table_name, 'public')
                schema_row.append(schema)
            ws.append(schema_row)
            
            # 두 번째 행: 테이블 헤더
            header_row = ['Package', 'Class']
            for table_name in table_names:
                header_row.append(table_name)
            ws.append(header_row)
            
            # 데이터 행들
            for _, row in df.iterrows():
                data_row = [row['Package'], row['Class']]
                for table_name in table_names:
                    operations = row[table_name]
                    data_row.append(operations)
                ws.append(data_row)
            
            # 스타일 적용
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            class_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 첫 번째 행 (스키마 정보) 스타일 적용 - 헤더와 동일하게
            for col in range(1, len(table_names) + 3):  # Package + Class + 테이블 컬럼들
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 두 번째 행 (헤더) 스타일 적용
            for col in range(1, len(table_names) + 3):
                cell = ws.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 데이터 행 스타일 적용
            for row in range(3, len(class_names) + 3):  # 데이터는 3행부터 시작
                for col in range(1, len(table_names) + 3):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    if col <= 2:  # Package와 Class 컬럼
                        cell.fill = class_fill
                        cell.font = Font(bold=False)  # 보통글씨
                        cell.alignment = left_alignment  # 좌측정렬
                    else:  # 테이블 컬럼들
                        cell.alignment = center_alignment
                        # CRUD 연산이 있는 경우 굵은글씨로
                        if cell.value and cell.value != '-':
                            cell.font = Font(bold=True)
                        else:
                            cell.font = Font(bold=False)
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)  # 최대 20자로 제한
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 요약 정보 시트 추가
            summary_ws = wb.create_sheet("Summary")
            summary_data = [
                ["CRUD Matrix Summary", ""],
                ["Project Name", project_name],
                ["Total Classes", summary['total_classes']],
                ["Total Tables", summary['total_tables']],
                ["", ""],
                ["CRUD Operations", ""],
                ["Create (C)", summary['crud_stats']['C']],
                ["Read (R)", summary['crud_stats']['R']],
                ["Update (U)", summary['crud_stats']['U']],
                ["Delete (D)", summary['crud_stats']['D']],
                ["Other (O)", summary['crud_stats']['O']],
                ["", ""],
                ["Most Active Class", summary.get('most_active_class', 'N/A')],
                ["Most Used Table", summary.get('most_used_table', 'N/A')]
            ]
            
            for row_data in summary_data:
                summary_ws.append(row_data)
            
            # 요약 시트 스타일 적용
            for row in range(1, len(summary_data) + 1):
                for col in range(1, 3):
                    cell = summary_ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if row == 1:  # 제목 행
                        cell.font = Font(bold=True, size=14)
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    elif col == 1:  # 첫 번째 컬럼 (라벨)
                        cell.font = Font(bold=True)
                        cell.fill = class_fill
            
            # 컬럼 너비 조정
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 15
            
            # 파일 저장
            wb.save(output_file)
            self.logger.info(f"Excel 파일이 생성되었습니다: {output_file}")
            return True
            
        except ImportError as e:
            self.logger.error(f"필요한 라이브러리가 설치되지 않았습니다: {e}")
            self.logger.error("다음 명령어로 설치하세요: pip install pandas openpyxl")
            return False
        except Exception as e:
            self.logger.error(f"Excel 파일 생성 오류: {e}")
            return False
